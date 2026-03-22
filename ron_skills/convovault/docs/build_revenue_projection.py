"""
ConvoVault Revenue Projection Builder
Generates a public-facing Excel revenue model for Labyrinth Analytics Consulting.
Usage: python build_revenue_projection.py
Output: ConvoVault_Revenue_Projection.xlsx in the same directory
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

# ---- Color palette (dark analytics theme) ----
COLOR_DARK_BG   = "1A1A2E"
COLOR_ACCENT1   = "16213E"
COLOR_ACCENT2   = "0F3460"
COLOR_HIGHLIGHT = "E94560"
COLOR_WHITE     = "FFFFFF"
COLOR_LIGHT_GRAY= "F2F2F7"
COLOR_MID_GRAY  = "CCCCCC"
COLOR_GREEN     = "2ECC71"
COLOR_GOLD      = "F39C12"
COLOR_BLUE      = "3498DB"

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color=COLOR_DARK_BG, italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)

def normal_font(size=11, color=COLOR_DARK_BG, bold=False):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def thin_border():
    side = Side(style="thin", color=COLOR_MID_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def right_align():
    return Alignment(horizontal="right", vertical="center")


def apply_header_row(ws, row, col_start, col_end, text, bg=COLOR_DARK_BG, fg=COLOR_WHITE, font_size=13):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = bold_font(size=font_size, color=fg)
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---- Sheet 1: Cover ----
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 2

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 25
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 20

    # Title block
    ws.merge_cells("B2:B2")
    c = ws["B2"]
    c.value = "ConvoVault"
    c.font = bold_font(size=36, color=COLOR_HIGHLIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws["B3"].value = "Revenue Projection Model -- 12-Month Outlook"
    ws["B3"].font = bold_font(size=16, color=COLOR_ACCENT2)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    ws["B4"].value = "Labyrinth Analytics Consulting"
    ws["B4"].font = normal_font(size=13, color=COLOR_MID_GRAY, bold=False)

    ws["B6"].value = "Prepared by: Ron (AI Product Agent)"
    ws["B6"].font = normal_font(size=11, color=COLOR_DARK_BG)

    ws["B7"].value = "Version: 1.0  |  Date: March 2026  |  Status: DRAFT -- Debbie must approve before sharing externally"
    ws["B7"].font = normal_font(size=10, color=COLOR_HIGHLIGHT, bold=False)

    ws["B9"].value = "Contents:"
    ws["B9"].font = bold_font(size=11, color=COLOR_DARK_BG)
    for i, item in enumerate(["Cover", "Pricing & Tiers", "Monthly Revenue Model", "Assumptions", "MRR Ramp Chart"], start=1):
        ws.cell(row=9 + i, column=2).value = "  " + item
        ws.cell(row=9 + i, column=2).font = normal_font(size=11)

    return ws


# ---- Sheet 2: Pricing Tiers ----
def build_pricing(wb):
    ws = wb.create_sheet("Pricing & Tiers")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 28, 16, 16, 16, 2])

    apply_header_row(ws, 1, 2, 5, "ConvoVault -- Pricing Tiers", COLOR_DARK_BG)
    ws.row_dimensions[1].height = 32

    headers = ["Feature", "Free", "Pro ($8/mo)", "Team ($19/mo)"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bold_font(size=11, color=COLOR_WHITE)
        c.fill = cell_fill(COLOR_ACCENT2)
        c.alignment = center()
        c.border = thin_border()

    rows = [
        ("Sessions stored",         "50",           "Unlimited",    "Unlimited"),
        ("Projects",                "3",            "Unlimited",    "Unlimited"),
        ("Full-text search",        "Yes",          "Yes",          "Yes"),
        ("Persona tagging",         "Yes",          "Yes",          "Yes"),
        ("Session linking",         "No",           "Yes",          "Yes"),
        ("vault_suggest (proactive)","No",           "Yes",          "Yes"),
        ("Chat bridge export",      "Yes",          "Yes",          "Yes"),
        ("Cloud sync (coming)",     "No",           "No",           "Yes"),
        ("Team shared vault (coming)","No",         "No",           "Yes"),
        ("Priority support",        "No",           "No",           "Yes"),
    ]

    for i, (feature, free, pro, team) in enumerate(rows):
        row_idx = 4 + i
        ws.row_dimensions[row_idx].height = 20
        bg = COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE
        vals = [feature, free, pro, team]
        for col, val in enumerate(vals, start=2):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = cell_fill(bg)
            c.border = thin_border()
            c.alignment = center() if col > 2 else Alignment(horizontal="left", vertical="center")
            if val in ("Yes",):
                c.font = normal_font(color=COLOR_GREEN, bold=True)
            elif val in ("No",):
                c.font = normal_font(color=COLOR_MID_GRAY)
            elif val == "Unlimited":
                c.font = normal_font(color=COLOR_BLUE, bold=True)
            else:
                c.font = normal_font()

    # Price row at top
    ws.row_dimensions[15].height = 26
    for col, (label, color) in enumerate(
        [("Free", COLOR_DARK_BG), ("$8/month", COLOR_ACCENT2), ("$19/month", COLOR_ACCENT2)], start=3
    ):
        c = ws.cell(row=15, column=col, value=label)
        c.font = bold_font(size=13, color=COLOR_WHITE)
        c.fill = cell_fill(color)
        c.alignment = center()
        c.border = thin_border()

    ws.cell(row=15, column=2, value="Monthly Price").font = bold_font(size=11, color=COLOR_DARK_BG)
    ws.cell(row=15, column=2).alignment = Alignment(horizontal="left", vertical="center")

    return ws


# ---- Sheet 3: Monthly Revenue Model ----
def build_revenue_model(wb):
    ws = wb.create_sheet("Monthly Revenue Model")
    ws.sheet_view.showGridLines = False

    months = [
        "Month 1", "Month 2", "Month 3", "Month 4", "Month 5", "Month 6",
        "Month 7", "Month 8", "Month 9", "Month 10", "Month 11", "Month 12"
    ]

    # Customer ramp: growth curve targeting $3,268 MRR by Month 12
    # Month 12: ~310 Pro + ~42 Team = $2,480 + $798 = $3,278 MRR
    free_users      = [  30,  70, 130, 210, 310, 430, 570, 720, 880,1040,1200,1380]
    pro_users       = [   3,   8,  17,  30,  50,  75, 110, 150, 195, 245, 278, 310]
    team_users      = [   0,   0,   1,   2,   4,   7,  11,  16,  22,  30,  37,  42]

    pro_price  = 8
    team_price = 19

    set_col_widths(ws, [2, 18] + [12] * 12 + [2])
    apply_header_row(ws, 1, 2, 14, "ConvoVault -- 12-Month Revenue Model", COLOR_DARK_BG)
    ws.row_dimensions[1].height = 32

    # Column headers
    ws.cell(row=3, column=2, value="Metric").font = bold_font(color=COLOR_WHITE)
    ws.cell(row=3, column=2).fill = cell_fill(COLOR_ACCENT2)
    ws.cell(row=3, column=2).border = thin_border()
    for i, m in enumerate(months):
        c = ws.cell(row=3, column=3 + i, value=m)
        c.font = bold_font(size=10, color=COLOR_WHITE)
        c.fill = cell_fill(COLOR_ACCENT2)
        c.alignment = center()
        c.border = thin_border()

    row_data = [
        ("--- Users ---",     None, None, None,        COLOR_ACCENT1, COLOR_WHITE, True),
        ("Free Users",        free_users, None, None,  None, None, False),
        ("Pro Users",         pro_users, None, None,   None, None, False),
        ("Team Users",        team_users, None, None,  None, None, False),
        ("--- Revenue ---",   None, None, None,        COLOR_ACCENT1, COLOR_WHITE, True),
        ("Pro MRR ($)",       None, pro_users, pro_price,  None, None, False),
        ("Team MRR ($)",      None, team_users, team_price, None, None, False),
        ("Total MRR ($)",     None, None, "sum",       None, None, False),
        ("--- Growth ---",    None, None, None,        COLOR_ACCENT1, COLOR_WHITE, True),
        ("MoM Growth (%)",    None, None, "growth",    None, None, False),
        ("Cumulative ARR ($)",None, None, "arr",       None, None, False),
    ]

    mrr_row_idx  = None
    pro_row_idx  = None
    team_row_idx = None

    for row_offset, (label, direct, user_list, modifier, bg_override, fg_override, is_section) in enumerate(row_data):
        row_idx = 4 + row_offset
        ws.row_dimensions[row_idx].height = 20

        lc = ws.cell(row=row_idx, column=2, value=label)
        if is_section:
            lc.font = bold_font(size=10, color=fg_override or COLOR_WHITE)
            lc.fill = cell_fill(bg_override or COLOR_ACCENT1)
        else:
            lc.font = bold_font(size=10, color=COLOR_DARK_BG)
            lc.fill = cell_fill(COLOR_LIGHT_GRAY)
        lc.border = thin_border()
        lc.alignment = Alignment(horizontal="left", vertical="center")

        for col_i in range(12):
            col_idx = 3 + col_i
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = thin_border()
            c.alignment = center()

            if is_section:
                c.fill = cell_fill(bg_override or COLOR_ACCENT1)
                continue

            bg = COLOR_LIGHT_GRAY if col_i % 2 == 0 else COLOR_WHITE
            c.fill = cell_fill(bg)

            if direct is not None:
                c.value = direct[col_i]
                c.font = normal_font(size=10)
            elif modifier == "sum" and pro_row_idx and team_row_idx:
                pro_ref  = f"{get_column_letter(col_idx)}{pro_row_idx}"
                team_ref = f"{get_column_letter(col_idx)}{team_row_idx}"
                c.value = f"={pro_ref}+{team_ref}"
                c.font  = bold_font(size=10, color=COLOR_HIGHLIGHT)
                if mrr_row_idx is None:
                    mrr_row_idx = row_idx
            elif modifier == "growth":
                if col_i == 0:
                    c.value = "N/A"
                    c.font  = normal_font(size=10, color=COLOR_MID_GRAY)
                else:
                    prev_col = get_column_letter(col_idx - 1)
                    curr_col = get_column_letter(col_idx)
                    c.value  = f"=IF({prev_col}{mrr_row_idx}=0,\"N/A\",({curr_col}{mrr_row_idx}-{prev_col}{mrr_row_idx})/{prev_col}{mrr_row_idx})"
                    c.number_format = "0.0%"
                    c.font = normal_font(size=10, color=COLOR_BLUE)
            elif modifier == "arr" and mrr_row_idx:
                curr_col = get_column_letter(col_idx)
                c.value  = f"={curr_col}{mrr_row_idx}*12"
                c.number_format = '"$"#,##0'
                c.font   = bold_font(size=10, color=COLOR_GREEN)
            elif user_list is not None and modifier is not None:
                c.value = user_list[col_i] * modifier
                c.number_format = '"$"#,##0'
                c.font  = normal_font(size=10)
                if label.startswith("Pro"):
                    pro_row_idx = row_idx
                if label.startswith("Team"):
                    team_row_idx = row_idx

    # Final MRR highlight note
    ws.row_dimensions[16].height = 24
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=14)
    note = ws.cell(row=16, column=2)
    note.value = "Target: $3,268 MRR by Month 12  |  Based on conservative user growth assumptions  |  See Assumptions sheet for full details"
    note.font  = bold_font(size=10, color=COLOR_WHITE)
    note.fill  = cell_fill(COLOR_ACCENT2)
    note.alignment = Alignment(horizontal="center", vertical="center")

    return ws, mrr_row_idx


# ---- Sheet 4: Assumptions ----
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 36, 30, 2])

    apply_header_row(ws, 1, 2, 3, "Model Assumptions", COLOR_DARK_BG)
    ws.row_dimensions[1].height = 32

    for col, hdr in enumerate(["Assumption", "Value / Notes"], start=2):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font = bold_font(color=COLOR_WHITE)
        c.fill = cell_fill(COLOR_ACCENT2)
        c.border = thin_border()
        c.alignment = center()

    assumptions = [
        ("--- Pricing ---",              ""),
        ("Pro tier price",               "$8/month per user"),
        ("Team tier price",              "$19/month per user"),
        ("Free tier",                    "No revenue; drives top-of-funnel"),
        ("--- Conversion ---",           ""),
        ("Free-to-Pro conversion",       "~16% over 12 months (conservative)"),
        ("Free-to-Team conversion",      "~3% over 12 months"),
        ("Monthly churn (Pro)",          "~5% assumed (not modeled -- use as buffer)"),
        ("--- Growth ---",               ""),
        ("Month 1 installs",             "20 Free / 2 Pro / 0 Team (early adopters)"),
        ("Growth driver",                "Organic: Claude Code community, GitHub, word of mouth"),
        ("Month 12 target",              "1,380 Free / 310 Pro / 42 Team"),
        ("--- Revenue ---",              ""),
        ("Month 12 MRR target",          "$3,268"),
        ("Month 12 ARR target",          "$39,216"),
        ("Time to first $1K MRR",        "~Month 8 (based on model)"),
        ("--- Distribution ---",         ""),
        ("Primary channel",              "Claude plugin marketplace (ClawHub / Salable)"),
        ("Secondary channel",            "GitHub open-source (drives free users)"),
        ("Billing platform",             "Salable (subscription gating)"),
        ("--- Risk Factors ---",         ""),
        ("Upside risk",                  "Viral adoption via Claude ecosystem growth"),
        ("Downside risk",                "Low free-to-paid conversion if hooks not compelling"),
        ("Mitigation",                   "vault_suggest, session linking are key Pro hooks"),
    ]

    for i, (assumption, value) in enumerate(assumptions):
        row_idx = 4 + i
        ws.row_dimensions[row_idx].height = 20
        is_section = assumption.startswith("---")
        bg = COLOR_ACCENT1 if is_section else (COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE)
        fg = COLOR_WHITE if is_section else COLOR_DARK_BG

        for col, val in enumerate([assumption, value], start=2):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = bold_font(size=10, color=fg) if is_section else normal_font(size=10, color=fg)
            c.fill = cell_fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")

    return ws


# ---- Sheet 5: Chart data + chart ----
def build_chart_sheet(wb, mrr_values):
    ws = wb.create_sheet("MRR Ramp Chart")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 14, 16, 2])

    apply_header_row(ws, 1, 2, 3, "MRR Ramp -- 12 Month Projection", COLOR_DARK_BG)
    ws.row_dimensions[1].height = 32

    ws.cell(row=3, column=2, value="Month").font = bold_font(color=COLOR_WHITE)
    ws.cell(row=3, column=2).fill = cell_fill(COLOR_ACCENT2)
    ws.cell(row=3, column=2).border = thin_border()
    ws.cell(row=3, column=3, value="Total MRR ($)").font = bold_font(color=COLOR_WHITE)
    ws.cell(row=3, column=3).fill = cell_fill(COLOR_ACCENT2)
    ws.cell(row=3, column=3).border = thin_border()

    for i, (m, mrr) in enumerate(zip(
        [f"Month {j}" for j in range(1, 13)],
        mrr_values
    )):
        row_idx = 4 + i
        ws.row_dimensions[row_idx].height = 18
        bg = COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE
        c1 = ws.cell(row=row_idx, column=2, value=m)
        c1.font = normal_font(size=10)
        c1.fill = cell_fill(bg)
        c1.border = thin_border()
        c2 = ws.cell(row=row_idx, column=3, value=mrr)
        c2.font = normal_font(size=10)
        c2.number_format = '"$"#,##0'
        c2.fill = cell_fill(bg)
        c2.border = thin_border()

    # Add line chart
    chart = LineChart()
    chart.title = "ConvoVault MRR Growth Projection"
    chart.style  = 10
    chart.y_axis.title = "Monthly Recurring Revenue ($)"
    chart.x_axis.title = "Month"
    chart.y_axis.numFmt = '"$"#,##0'
    chart.height = 15
    chart.width  = 25

    data_ref = Reference(ws, min_col=3, min_row=3, max_row=15)
    cats_ref = Reference(ws, min_col=2, min_row=4, max_row=15)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = COLOR_HIGHLIGHT
    chart.series[0].graphicalProperties.line.width = 25000  # in EMU

    ws.add_chart(chart, "E3")
    return ws


def main():
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "ConvoVault_Revenue_Projection.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb)
    build_pricing(wb)

    # Revenue model: compute actual MRR values for chart
    pro_users  = [  3,   8,  17,  30,  50,  75, 110, 150, 195, 245, 278, 310]
    team_users = [  0,   0,   1,   2,   4,   7,  11,  16,  22,  30,  37,  42]
    mrr_values = [p * 8 + t * 19 for p, t in zip(pro_users, team_users)]

    ws_model, mrr_row = build_revenue_model(wb)
    build_assumptions(wb)
    build_chart_sheet(wb, mrr_values)

    wb.save(out_path)
    print(f"[OK] Saved: {out_path}")
    print(f"[OK] Month 12 MRR: ${mrr_values[-1]:,}")
    print(f"[OK] Month 12 ARR: ${mrr_values[-1]*12:,}")


if __name__ == "__main__":
    main()
